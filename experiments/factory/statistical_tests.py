"""
Statistical Tests for Neuromorphic MRTA Paper
===============================================
Runs Wilcoxon signed-rank, Mann-Whitney U, and effect size (Cohen's d)
comparing OIM vs Greedy, SNN vs Greedy, OIM vs SA across factory scales.

Requires: experiments/datasets/factory_benchmarks.xlsx (run benchmarks first)
Outputs:  experiments/datasets/statistical_tests.xlsx
"""
from __future__ import annotations
import sys, warnings
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from scipy import stats
except ImportError:
    print("scipy required: pip install scipy")
    sys.exit(1)

warnings.filterwarnings("ignore")

SCALE_SHEETS = {
    "Small (3R5T)": "Small",
    "Medium (5R8T)": "Medium",
    "Large (7R10T)": "Large",
    "Mega (10R12T)": "Mega",
}


def cohens_d(a: np.ndarray, b: np.ndarray) -> float:
    """Compute Cohen's d effect size. Returns 0 when both arrays have zero variance."""
    n_a, n_b = len(a), len(b)
    pooled_var = ((n_a-1)*np.var(a, ddof=1) + (n_b-1)*np.var(b, ddof=1)) / (n_a+n_b-2)
    if pooled_var < 1e-12:
        # Both distributions are deterministic; return 0 if equal, cap at ±4 if different
        diff = np.mean(a) - np.mean(b)
        scale = max(abs(np.mean(a)), abs(np.mean(b)), 1e-12)
        return np.clip(diff / scale * 4.0, -4.0, 4.0)
    return (np.mean(a) - np.mean(b)) / np.sqrt(pooled_var)


def interpret_d(d: float) -> str:
    ad = abs(d)
    if ad < 0.2: return "negligible"
    if ad < 0.5: return "small"
    if ad < 0.8: return "medium"
    return "large"


def interpret_p(p: float) -> str:
    if p < 0.001: return "***"
    if p < 0.01:  return "**"
    if p < 0.05:  return "*"
    return "n.s."


def run_tests(df: pd.DataFrame, scale_label: str) -> list[dict]:
    """Run all pairwise statistical tests for a single factory scale."""
    results = []
    pairs = [("oim", "greedy"), ("snn", "greedy"), ("oim", "sa"), ("snn", "sa"), ("oim", "snn")]

    for a_solver, b_solver in pairs:
        a_col = f"{a_solver}_utility"
        b_col = f"{b_solver}_utility"
        if a_col not in df.columns or b_col not in df.columns:
            continue
        a = df[a_col].dropna().values
        b = df[b_col].dropna().values
        n = min(len(a), len(b))
        a, b = a[:n], b[:n]
        if n < 5:
            continue

        # Wilcoxon signed-rank (paired)
        try:
            w_stat, w_p = stats.wilcoxon(a, b, alternative="two-sided")
        except Exception:
            w_stat, w_p = np.nan, np.nan

        # Mann-Whitney U (unpaired)
        u_stat, u_p = stats.mannwhitneyu(a, b, alternative="two-sided")

        # Effect size
        d = cohens_d(a, b)

        results.append({
            "scale": scale_label,
            "comparison": f"{a_solver.upper()} vs {b_solver.upper()}",
            "n_trials": n,
            "mean_A": round(np.mean(a), 4),
            "mean_B": round(np.mean(b), 4),
            "diff_mean": round(np.mean(a) - np.mean(b), 4),
            "diff_pct": round((np.mean(a)-np.mean(b))/np.mean(b)*100 if np.mean(b)>0 else 0, 2),
            "wilcoxon_stat": round(w_stat, 3) if not np.isnan(w_stat) else "n/a",
            "wilcoxon_p": round(w_p, 5) if not np.isnan(w_p) else "n/a",
            "wilcoxon_sig": interpret_p(w_p) if not np.isnan(w_p) else "n/a",
            "mannwhitney_U": round(u_stat, 1),
            "mannwhitney_p": round(u_p, 5),
            "mannwhitney_sig": interpret_p(u_p),
            "cohens_d": round(d, 4),
            "effect_size": interpret_d(d),
            "A_wins_pct": round(np.mean(a > b) * 100, 1),
        })
    return results


def main():
    bench_path = Path(__file__).parent.parent / "datasets" / "factory_benchmarks.xlsx"
    if not bench_path.exists():
        print(f"ERROR: {bench_path} not found. Run run_factory_benchmarks.py first.")
        sys.exit(1)

    all_results = []
    print("Statistical Tests — Neuromorphic MRTA")
    print("=" * 75)

    xl = pd.ExcelFile(bench_path)
    available = xl.sheet_names

    scale_map = {}
    for label, sheet_prefix in SCALE_SHEETS.items():
        # Find matching sheet
        match = next((s for s in available if s.startswith(sheet_prefix)), None)
        if match:
            scale_map[label] = match

    if not scale_map:
        print("No matching sheets found. Available:", available)
        sys.exit(1)

    for scale_label, sheet_name in scale_map.items():
        df = pd.read_excel(bench_path, sheet_name=sheet_name)
        tests = run_tests(df, scale_label)
        all_results.extend(tests)
        print(f"\n{scale_label}:")
        print(f"  {'Comparison':<22} {'diff%':>7} {'W_p':>8} {'Sig':>5} {'Cohen_d':>8} {'Effect':>10} {'A_wins%':>8}")
        for t in tests:
            print(f"  {t['comparison']:<22} {t['diff_pct']:>7.1f} {str(t['wilcoxon_p']):>8} {t['wilcoxon_sig']:>5} {t['cohens_d']:>8.3f} {t['effect_size']:>10} {t['A_wins_pct']:>8.1f}")

    # Scaling study tests
    if "ScalingStudy" in available:
        df_sc = pd.read_excel(bench_path, sheet_name="ScalingStudy")
        print("\nScaling Study — OIM vs Greedy by graph size:")
        print(f"  {'n_nodes':>8} {'OIM_mean':>10} {'Gr_mean':>10} {'diff%':>7} {'p':>8} {'sig':>5}")
        for n_nodes in sorted(df_sc["n_nodes"].unique()):
            sub = df_sc[df_sc["n_nodes"] == n_nodes]
            a = sub["oim_utility"].values
            b = sub["greedy_utility"].values
            try:
                _, p = stats.wilcoxon(a, b)
            except Exception:
                p = np.nan
            diff_pct = (np.mean(a)-np.mean(b))/np.mean(b)*100 if np.mean(b)>0 else 0
            sig = interpret_p(p) if not np.isnan(p) else "n/a"
            print(f"  {n_nodes:>8}  {np.mean(a):>10.3f}  {np.mean(b):>10.3f}  {diff_pct:>7.1f}  {p if not np.isnan(p) else 'n/a':>8}  {sig:>5}")

    # ── Save Excel ─────────────────────────────────────────────────────────
    out_path = Path(__file__).parent.parent / "datasets" / "statistical_tests.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistical_Tests"
    if not all_results:
        ws.append(["No data — run benchmarks first"])
        wb.save(out_path)
        return

    headers = list(all_results[0].keys())
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in all_results:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            16, len(str(col[0].value or "")) + 2)

    # Significance guide sheet
    ws2 = wb.create_sheet("Legend")
    for row in [
        ["Symbol", "Meaning"],
        ["***", "p < 0.001 (highly significant)"],
        ["**", "p < 0.01 (significant)"],
        ["*", "p < 0.05 (marginally significant)"],
        ["n.s.", "p ≥ 0.05 (not significant)"],
        ["", ""],
        ["Effect size (Cohen's d):", ""],
        ["< 0.2", "Negligible"],
        ["0.2 – 0.5", "Small"],
        ["0.5 – 0.8", "Medium"],
        ["> 0.8", "Large"],
    ]:
        ws2.append(row)

    wb.save(out_path)
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()
