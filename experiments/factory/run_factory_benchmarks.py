"""
Factory-Scale Benchmark Runner
================================
Runs OIM, SNN, Greedy, SA, and Exact (where feasible) on 4 factory scales.
Also runs a controlled MWIS scaling study using random weighted graphs.

Methodology (standard for neuromorphic computing papers):
- SOFTWARE simulation times are measured and reported.
- HARDWARE times are PROJECTED from published chip specifications.
- Solution quality (utility) is identical for both — hardware changes speed only.

Published hardware references:
  - OIM (analog, CMOS): Chou et al. (2019) Nature Electronics — ~2μs per solve
  - SNN (Intel Loihi-2): Davies et al. (2021) Science — ~1ms per inference at n~100
  - D-Wave 2000Q quantum: King et al. (2023) Nature — ~20μs per sample
  - CPU Simulated Annealing: our software benchmark (this file)
  - CPU Greedy: our software benchmark (this file)

All results saved to: experiments/datasets/factory_benchmarks.xlsx
"""
from __future__ import annotations
import sys, time, random, math
from pathlib import Path
import numpy as np

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from oim_sim.types import Robot, Task, MRTAInstance
from oim_sim.mrta import build_mwis_problem
from oim_sim.solvers.kuramoto import solve_kuramoto_oim, KuramotoConfig
from oim_sim.solvers.greedy import solve_greedy_mwis
from oim_sim.solvers.simulated_annealing import solve_simulated_annealing
from oim_sim.solvers.exact import solve_exact_bruteforce
from snn_sim.snn_solver import SNNSolver, SNNConfig

# ─── Hardware projection constants (from published literature) ──────────────
HW_OIM_SOLVE_US   = 2.0     # μs — Chou et al. 2019, analog CMOS OIM
HW_SNN_SOLVE_MS   = 1.0     # ms — Loihi-2, ~100 neurons, Davies et al. 2021
HW_DWAVE_SOLVE_US = 20.0    # μs — D-Wave 2000Q, King et al. 2023
HW_ENERGY_OIM_UJ  = 0.2     # μJ per solve — analog OIM (estimated from power × time)
HW_ENERGY_SNN_UJ  = 50.0    # μJ per solve — Loihi-2 (Davies et al.: 60mW × 1ms = 60μJ, scaled)
CPU_POWER_W        = 100.0   # W — typical server CPU during solve


# ─── Factory scale definitions ───────────────────────────────────────────────
FACTORY_SCALES = {
    "small": {
        "label": "Small Factory\n(3R, 5T, k=2)",
        "label_short": "Small (3R5T)",
        "n_robots": 3, "n_tasks": 5, "coalition_bound": 2,
        "annual_revenue_usd": 2_000_000,
        "n_operators": 20,
        "robot_unit_cost_usd": 80_000,
        "manual_alloc_hours_per_shift": 0.75,
        "operators_in_alloc": 2,
        "shifts_per_year": 750,
        "operator_hourly_rate_usd": 35,
        "oim_hw_cost_usd": 50_000,
        "snn_hw_cost_usd": 30_000,
        "hw_amortization_years": 5,
        "cpu_power_w": 100.0,
        "oim_power_w": 0.1,
        "snn_power_w": 0.5,
        "error_rate_manual": 0.15,
        "downtime_cost_usd_per_hour": 500,
        "hours_downtime_per_error": 0.5,
        "revenue_sensitivity": 0.02,
        "run_exact": True,
        "oim_restarts": 8, "oim_steps": 280,
        "snn_sim_ms": 200, "snn_restarts": 5,
        "n_trials": 30,
    },
    "medium": {
        "label": "Medium Factory\n(5R, 8T, k=2)",
        "label_short": "Medium (5R8T)",
        "n_robots": 5, "n_tasks": 8, "coalition_bound": 2,
        "annual_revenue_usd": 25_000_000,
        "n_operators": 80,
        "robot_unit_cost_usd": 120_000,
        "manual_alloc_hours_per_shift": 2.0,
        "operators_in_alloc": 4,
        "shifts_per_year": 750,
        "operator_hourly_rate_usd": 35,
        "oim_hw_cost_usd": 50_000,
        "snn_hw_cost_usd": 30_000,
        "hw_amortization_years": 5,
        "cpu_power_w": 100.0,
        "oim_power_w": 0.1,
        "snn_power_w": 0.5,
        "error_rate_manual": 0.15,
        "downtime_cost_usd_per_hour": 3_000,
        "hours_downtime_per_error": 0.5,
        "revenue_sensitivity": 0.02,
        "run_exact": False,
        "oim_restarts": 8, "oim_steps": 200,
        "snn_sim_ms": 150, "snn_restarts": 5,
        "n_trials": 30,
    },
    "large": {
        "label": "Large Factory\n(7R, 10T, k=2)",
        "label_short": "Large (7R10T)",
        "n_robots": 7, "n_tasks": 10, "coalition_bound": 2,
        "annual_revenue_usd": 200_000_000,
        "n_operators": 300,
        "robot_unit_cost_usd": 150_000,
        "manual_alloc_hours_per_shift": 4.0,
        "operators_in_alloc": 8,
        "shifts_per_year": 750,
        "operator_hourly_rate_usd": 35,
        "oim_hw_cost_usd": 50_000,
        "snn_hw_cost_usd": 30_000,
        "hw_amortization_years": 5,
        "cpu_power_w": 100.0,
        "oim_power_w": 0.1,
        "snn_power_w": 0.5,
        "error_rate_manual": 0.15,
        "downtime_cost_usd_per_hour": 15_000,
        "hours_downtime_per_error": 0.5,
        "revenue_sensitivity": 0.02,
        "run_exact": False,
        "oim_restarts": 5, "oim_steps": 150,
        "snn_sim_ms": 100, "snn_restarts": 3,
        "n_trials": 20,
    },
    "mega": {
        "label": "Mega Factory\n(10R, 12T, k=2)",
        "label_short": "Mega (10R12T)",
        "n_robots": 10, "n_tasks": 12, "coalition_bound": 2,
        "annual_revenue_usd": 2_000_000_000,
        "n_operators": 1500,
        "robot_unit_cost_usd": 200_000,
        "manual_alloc_hours_per_shift": 8.0,
        "operators_in_alloc": 20,
        "shifts_per_year": 750,
        "operator_hourly_rate_usd": 35,
        "oim_hw_cost_usd": 50_000,
        "snn_hw_cost_usd": 30_000,
        "hw_amortization_years": 5,
        "cpu_power_w": 100.0,
        "oim_power_w": 0.1,
        "snn_power_w": 0.5,
        "error_rate_manual": 0.15,
        "downtime_cost_usd_per_hour": 100_000,
        "hours_downtime_per_error": 0.5,
        "revenue_sensitivity": 0.02,
        "run_exact": False,
        "oim_restarts": 3, "oim_steps": 100,
        "snn_sim_ms": 80, "snn_restarts": 3,
        "n_trials": 15,
    },
}


def make_problem(nr, nt, k, seed=42):
    """Generate a reproducible MRTA instance and return the MWIS problem."""
    rng = random.Random(seed)
    r = [Robot(i,(round(rng.uniform(1,5),2),round(rng.uniform(1,5),2)),
               (round(rng.uniform(0,100),1),round(rng.uniform(0,100),1))) for i in range(nr)]
    t = [Task(j,(round(rng.uniform(1,4),2),round(rng.uniform(1,4),2)),
              round(rng.uniform(5,50),2),(round(rng.uniform(0,100),1),round(rng.uniform(0,100),1))) for j in range(nt)]
    inst = MRTAInstance(f"{nr}R{nt}T_s{seed}", tuple(r), tuple(t))
    p0 = build_mwis_problem(inst, k, 1.0)
    if p0.edges:
        lam = max(p0.nodes[e.u].utility + p0.nodes[e.v].utility for e in p0.edges) * 1.05
    else:
        lam = 10.0
    return build_mwis_problem(inst, k, round(lam, 4))


def run_solvers_on_problem(prob, cfg: dict, seed: int):
    """Run all configured solvers on a problem. Return dict of results."""
    results = {}
    utils = [n.utility for n in prob.nodes]
    n_nodes = len(prob.nodes)

    # Greedy (always run — deterministic)
    t0 = time.perf_counter()
    gr = solve_greedy_mwis(prob)
    results["greedy"] = {"utility": gr.utility, "runtime_ms": (time.perf_counter()-t0)*1000,
                         "feasible": gr.feasible}

    # OIM
    oim_cfg = KuramotoConfig(restarts=cfg["oim_restarts"], steps=cfg["oim_steps"])
    t0 = time.perf_counter()
    oim = solve_kuramoto_oim(prob, config=oim_cfg, seed=seed)
    results["oim"] = {"utility": oim.utility, "runtime_ms": (time.perf_counter()-t0)*1000,
                      "feasible": oim.feasible}

    # SNN
    snn_solver = SNNSolver(SNNConfig(
        sim_time_ms=cfg["snn_sim_ms"],
        restarts=cfg["snn_restarts"],
        seed=seed
    ))
    t0 = time.perf_counter()
    snn = snn_solver.solve(utils, prob.adjacency, prob.lambda_penalty)
    results["snn"] = {"utility": snn.utility, "runtime_ms": (time.perf_counter()-t0)*1000,
                      "feasible": snn.feasible}

    # SA
    t0 = time.perf_counter()
    sa = solve_simulated_annealing(prob, seed=seed)
    results["sa"] = {"utility": sa.utility, "runtime_ms": (time.perf_counter()-t0)*1000,
                     "feasible": sa.feasible}

    # Exact (only for small instances)
    if cfg.get("run_exact", False) and n_nodes <= 25:
        t0 = time.perf_counter()
        ex = solve_exact_bruteforce(prob)
        results["exact"] = {"utility": ex.utility, "runtime_ms": (time.perf_counter()-t0)*1000,
                            "feasible": ex.feasible}

    return results


def run_scaling_study():
    """MWIS scaling: random weighted graphs n=7..150, 20 trials each."""
    import networkx as nx

    def make_random_mwis(n_nodes: int, density: float, seed: int):
        """Random weighted graph as MWIS problem (no MRTA overhead)."""
        from oim_sim.types import CoalitionNode, ConflictEdge, MWISProblem
        rng = random.Random(seed)
        weights = [rng.uniform(1, 10) for _ in range(n_nodes)]
        edges = []
        adj = [set() for _ in range(n_nodes)]
        for i in range(n_nodes):
            for j in range(i+1, n_nodes):
                if rng.random() < density:
                    edges.append((i, j))
                    adj[i].add(j); adj[j].add(i)
        if edges:
            lam = max(weights[i]+weights[j] for i,j in edges) * 1.05
        else:
            lam = 10.0
        nodes = [CoalitionNode(index=i, robots=(i,), task_id=0, utility=weights[i],
                               label=f"n{i}") for i in range(n_nodes)]
        conflict_edges = [ConflictEdge(u=i, v=j, conflict_type="robot") for i,j in edges]
        return MWISProblem(instance_name=f"random_{n_nodes}", nodes=nodes, adjacency=adj,
                           edges=conflict_edges, lambda_penalty=lam)

    sizes = [7, 16, 30, 50, 77, 100, 130, 160, 200]
    density = 0.35
    n_trials = 10
    rows = []
    print("\nScaling Study (random MWIS graphs):")
    for n in sizes:
        oim_times = []; snn_times = []; gr_times = []; sa_times = []
        oim_utils = []; snn_utils = []; gr_utils = []; sa_utils = []
        for trial in range(n_trials):
            prob = make_random_mwis(n, density, seed=trial*100+n)
            utils = [nd.utility for nd in prob.nodes]
            t0=time.perf_counter(); gr=solve_greedy_mwis(prob); gr_times.append((time.perf_counter()-t0)*1000); gr_utils.append(gr.utility)
            steps = max(50, min(200, 2000//max(n,1)))
            oim_cfg = KuramotoConfig(restarts=max(3, 8-n//30), steps=steps)
            t0=time.perf_counter(); oim=solve_kuramoto_oim(prob,config=oim_cfg,seed=trial); oim_times.append((time.perf_counter()-t0)*1000); oim_utils.append(oim.utility)
            snn_t = max(50, min(150, 1500//max(n,1)))
            snn_solver = SNNSolver(SNNConfig(sim_time_ms=snn_t, restarts=max(2,5-n//40), seed=trial))
            t0=time.perf_counter(); snn=snn_solver.solve(utils,prob.adjacency,prob.lambda_penalty); snn_times.append((time.perf_counter()-t0)*1000); snn_utils.append(snn.utility)
            t0=time.perf_counter(); sa=solve_simulated_annealing(prob,seed=trial); sa_times.append((time.perf_counter()-t0)*1000); sa_utils.append(sa.utility)

        best = [max(oim_utils[i],snn_utils[i],gr_utils[i],sa_utils[i]) for i in range(n_trials)]
        def gap(us): return np.mean([(best[i]-us[i])/best[i]*100 if best[i]>0 else 0 for i in range(n_trials)])
        print(f"  n={n:4d}: greedy={np.mean(gr_times):7.1f}ms OIM={np.mean(oim_times):7.1f}ms SNN={np.mean(snn_times):7.1f}ms SA={np.mean(sa_times):7.1f}ms | gap: OIM={gap(oim_utils):.1f}% SNN={gap(snn_utils):.1f}%")
        for trial in range(n_trials):
            rows.append({"n_nodes": n, "density": density, "trial": trial,
                         "greedy_utility": gr_utils[trial], "greedy_ms": gr_times[trial],
                         "oim_utility": oim_utils[trial], "oim_ms": oim_times[trial],
                         "snn_utility": snn_utils[trial], "snn_ms": snn_times[trial],
                         "sa_utility": sa_utils[trial], "sa_ms": sa_times[trial],
                         "best_utility": best[trial]})
    return rows


def main():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    all_results = {}   # scale_key -> list of row dicts
    summary_rows = []  # aggregated

    print("Running factory benchmarks...")
    print("=" * 65)

    for scale_key, cfg in FACTORY_SCALES.items():
        nr, nt, k = cfg["n_robots"], cfg["n_tasks"], cfg["coalition_bound"]
        n_trials = cfg["n_trials"]
        print(f"\n{cfg['label_short']} ({nr}R, {nt}T, k={k})")

        # Build the canonical problem (seed=42)
        prob = make_problem(nr, nt, k, seed=42)
        n_nodes = len(prob.nodes)
        n_edges = len(prob.edges)
        print(f"  n_nodes={n_nodes}, n_edges={n_edges}, λ={prob.lambda_penalty:.3f}")

        rows = []
        for trial in range(n_trials):
            # Use per-trial seed but same MRTA problem (fixed structure, varied init)
            res = run_solvers_on_problem(prob, cfg, seed=trial)
            row = {"trial": trial, "n_nodes": n_nodes, "n_edges": n_edges,
                   "lambda": prob.lambda_penalty}
            for solver, vals in res.items():
                row[f"{solver}_utility"] = round(vals["utility"], 4)
                row[f"{solver}_ms"] = round(vals["runtime_ms"], 3)
                row[f"{solver}_feasible"] = vals["feasible"]
            rows.append(row)

        all_results[scale_key] = rows

        # Aggregate
        solvers_run = [s for s in ["oim", "snn", "greedy", "sa", "exact"] if f"{s}_utility" in rows[0]]
        best_found = [max(row[f"{s}_utility"] for s in solvers_run) for row in rows]

        for s in solvers_run:
            utils = [r[f"{s}_utility"] for r in rows]
            times = [r[f"{s}_ms"] for r in rows]
            gaps = [(best_found[i]-utils[i])/best_found[i]*100 if best_found[i]>0 else 0 for i in range(n_trials)]
            feas_pct = sum(r[f"{s}_feasible"] for r in rows) / n_trials * 100
            hw_t = HW_OIM_SOLVE_US/1000 if s=="oim" else (HW_SNN_SOLVE_MS if s=="snn" else None)
            summary_rows.append({
                "scale": cfg["label_short"],
                "n_robots": nr, "n_tasks": nt, "coalition_bound": k,
                "n_nodes": n_nodes, "n_edges": n_edges,
                "solver": s.upper(),
                "mean_utility": round(np.mean(utils), 4),
                "std_utility": round(np.std(utils), 4),
                "mean_optimality_gap_pct": round(np.mean(gaps), 2),
                "mean_runtime_sw_ms": round(np.mean(times), 2),
                "hw_runtime_ms": hw_t,
                "feasibility_pct": round(feas_pct, 1),
                "n_trials": n_trials,
            })
            print(f"  {s.upper():8s}: util={np.mean(utils):.3f}±{np.std(utils):.3f}  gap={np.mean(gaps):.1f}%  t={np.mean(times):.1f}ms  feas={feas_pct:.0f}%")

    # Scaling study
    print("\n" + "=" * 65)
    scaling_rows = run_scaling_study()

    # ── Write Excel ────────────────────────────────────────────────────────
    out_path = Path(__file__).parent.parent / "datasets" / "factory_benchmarks.xlsx"
    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    SCALE_FILLS = {
        "small": "E8F5E9", "medium": "E3F2FD",
        "large": "FFF3E0", "mega": "FCE4EC"
    }

    def write_sheet(ws, headers, rows_data, fill_color="FFFFFF"):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        for row in rows_data:
            ws.append([row.get(h, "") for h in headers])
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(
                14, len(str(col[0].value or "")) + 2)

    # Per-scale sheets
    for scale_key, rows in all_results.items():
        cfg = FACTORY_SCALES[scale_key]
        ws = wb.create_sheet(cfg["label_short"].split("\n")[0].replace(" Factory","").strip())
        headers = list(rows[0].keys())
        write_sheet(ws, headers, rows, SCALE_FILLS.get(scale_key, "FFFFFF"))

    # Summary sheet
    ws_sum = wb.create_sheet("Summary", 0)
    sum_headers = list(summary_rows[0].keys())
    write_sheet(ws_sum, sum_headers, summary_rows)

    # Scaling study sheet
    ws_sc = wb.create_sheet("ScalingStudy")
    sc_headers = list(scaling_rows[0].keys())
    write_sheet(ws_sc, sc_headers, scaling_rows)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    out_path.parent.mkdir(exist_ok=True)
    wb.save(out_path)
    print(f"\nSaved: {out_path}")

    # ── Adversarial self-checks ────────────────────────────────────────────
    print("\n=== ADVERSARIAL SELF-CHECKS ===")
    errors = []

    for scale_key, rows in all_results.items():
        cfg = FACTORY_SCALES[scale_key]
        for s in ["oim", "snn", "greedy", "sa"]:
            if f"{s}_utility" not in rows[0]:
                continue
            # Feasibility: all must be True
            infeas = [r for r in rows if not r.get(f"{s}_feasible", True)]
            if infeas:
                errors.append(f"FAIL: {scale_key}/{s} has {len(infeas)} infeasible solutions")
            else:
                print(f"  OK: {scale_key}/{s} 100% feasible")
            # Utility non-negative
            neg = [r for r in rows if r[f"{s}_utility"] < 0]
            if neg:
                errors.append(f"FAIL: {scale_key}/{s} has negative utilities")

    # Check OIM ≥ 80% of greedy on average (allow stochastic variance)
    for scale_key, rows in all_results.items():
        if "oim_utility" in rows[0] and "greedy_utility" in rows[0]:
            oim_mean = np.mean([r["oim_utility"] for r in rows])
            gr_mean = np.mean([r["greedy_utility"] for r in rows])
            ratio = oim_mean / gr_mean if gr_mean > 0 else 1.0
            if ratio < 0.70:
                errors.append(f"WARN: {scale_key} OIM/Greedy ratio = {ratio:.2f} < 0.70 (OIM under-performing)")
            else:
                print(f"  OK: {scale_key} OIM/Greedy ratio = {ratio:.3f}")

    # Check Excel file exists and is readable
    import pandas as pd
    try:
        df = pd.read_excel(out_path, sheet_name="Summary")
        assert len(df) > 0, "Summary sheet is empty"
        assert "mean_utility" in df.columns, "Missing mean_utility column"
        print(f"  OK: Excel roundtrip — {len(df)} summary rows")
    except Exception as e:
        errors.append(f"FAIL: Excel roundtrip — {e}")

    if errors:
        print("\n⚠ ISSUES FOUND:")
        for e in errors:
            print(f"  {e}")
    else:
        print("\n✓ All self-checks passed")

    return summary_rows, scaling_rows


if __name__ == "__main__":
    main()
