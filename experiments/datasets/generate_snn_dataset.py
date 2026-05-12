"""Generate SNN Dataset Excel file with full traceability.

Produces: experiments/datasets/snn_2dof_dataset.xlsx
Sheets:
  1. ArmParams         - l1,l2,m1,m2,g,I values
  2. InertiaMatrix     - M(theta) at 3 configurations, step-by-step
  3. GravityTorques    - G(theta) at 3 configurations
  4. MRTA_Mapping      - coalition nodes -> neurons
  5. SNN_Dynamics      - voltage traces for all 7 neurons
  6. SpikeCounts       - final spike count per neuron, selected allocation
  7. Comparison        - SNN vs OIM on same 3R2T instance
"""
from __future__ import annotations

import math
import sys
import random
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from oim_sim.types import Robot, Task, MRTAInstance
from oim_sim.mrta import build_mwis_problem
from oim_sim.solvers.kuramoto import solve_kuramoto_oim
from snn_sim import SNNSolver, SNNConfig, ArmDynamics
from snn_sim.arm_dynamics import ArmParams

# ---- Style helpers ----
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(name="Calibri", italic=True, color="7F6000", size=9)
ALT_FILL = PatternFill("solid", fgColor="DEEAF1")
BODY_FONT = Font(name="Calibri", size=10)
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell, text, fill=None, font=None):
    cell.value = text
    cell.font = font or HDR_FONT
    cell.fill = fill or HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_body(cell, value, number_format=None, fill=None):
    cell.value = value
    cell.font = BODY_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---- Instances ----

def get_instance():
    robots = (
        Robot(id=0, capabilities=(2.0, 0.0), position=(0.0, 0.0)),
        Robot(id=1, capabilities=(0.0, 2.0), position=(1.0, 1.0)),
        Robot(id=2, capabilities=(1.0, 1.0), position=(2.0, 0.0)),
    )
    tasks = (
        Task(id=0, requirements=(1.0, 1.0), value=6.0, position=(0.5, 0.5)),
        Task(id=1, requirements=(2.0, 0.0), value=5.0, position=(2.0, 0.5)),
    )
    return MRTAInstance(name="3R2T_Worked_Example", robots=robots, tasks=tasks)


# ---- Sheet 1: ArmParams ----

def sheet_arm_params(wb):
    ws = wb.create_sheet("ArmParams")
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "2-DOF Planar Robot Arm — Physical Parameters"
    c.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    p = ArmParams()
    params = [
        ("l1", p.l1, "m", "Length of link 1"),
        ("l2", p.l2, "m", "Length of link 2"),
        ("m1", p.m1, "kg", "Mass of link 1"),
        ("m2", p.m2, "kg", "Mass of link 2"),
        ("g", p.g, "m/s^2", "Gravitational acceleration"),
        ("I1 = m1*l1^2/12", p.I1, "kg*m^2", "Moment of inertia, link 1 about its CoM"),
        ("I2 = m2*l2^2/12", p.I2, "kg*m^2", "Moment of inertia, link 2 about its CoM"),
        ("V_th (LIF threshold)", 1.0, "V", "LIF neuron firing threshold"),
        ("tau (membrane)", 20.0, "ms", "LIF membrane time constant"),
        ("tau_ref", 2.0, "ms", "LIF refractory period"),
        ("R (membrane resistance)", 1.0, "MOhm", "LIF membrane resistance"),
        ("lambda (QUBO penalty)", 8.0, "", "Conflict penalty coefficient"),
    ]

    headers = ["Parameter", "Value", "Unit", "Description"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(3, ci), h)

    for i, (name, val, unit, desc) in enumerate(params):
        row = 4 + i
        fill = ALT_FILL if i % 2 else None
        for ci, v in enumerate([name, val, unit, desc], 1):
            style_body(ws.cell(row, ci), v, fill=fill)

    ws["A17"].value = ("Note: I = ml^2/12 is moment of inertia of uniform rod about its center of mass "
                       "(parallel-axis theorem then gives the full matrix entries). "
                       "The thesis shorthand 'I=ml^2/3' refers to moment about the proximal joint end.")
    ws["A17"].font = NOTE_FONT; ws["A17"].fill = NOTE_FILL
    ws.merge_cells("A17:D17")

    set_col_widths(ws, {"A":30,"B":14,"C":14,"D":50})


# ---- Sheet 2: InertiaMatrix ----

def sheet_inertia_matrix(wb):
    ws = wb.create_sheet("InertiaMatrix")
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Inertia Matrix M(theta) — Step-by-Step Computation at 3 Configurations"
    c.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    p = ArmParams()
    arm = ArmDynamics(p)

    configs = [
        (0.0, 0.0, "theta=[0, 0]"),
        (math.pi/4, math.pi/4, "theta=[pi/4, pi/4]"),
        (math.pi/2, 0.0, "theta=[pi/2, 0]"),
    ]

    row = 3
    for t1, t2, name in configs:
        c2 = math.cos(t2)
        I1 = p.I1; I2 = p.I2

        M11 = I1 + p.m1*(p.l1/2)**2 + I2 + p.m2*(p.l1**2 + p.l1*p.l2*c2 + (p.l2/2)**2)
        M12 = I2 + p.m2*(p.l2/2)*(p.l1*c2 + p.l2/2)
        M22 = I2 + p.m2*(p.l2/2)**2

        # Title
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row, 1).value = name
        ws.cell(row, 1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row, 1).fill = SUBHDR_FILL
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        row += 1

        # Parameter values
        ws.cell(row, 1).value = "theta1 (rad)"
        ws.cell(row, 2).value = round(t1, 4)
        ws.cell(row, 3).value = "theta2 (rad)"
        ws.cell(row, 4).value = round(t2, 4)
        ws.cell(row, 5).value = "cos(theta2)"
        ws.cell(row, 6).value = round(c2, 6)
        for ci in range(1,7):
            ws.cell(row, ci).font = BODY_FONT
            ws.cell(row, ci).fill = NOTE_FILL
        row += 1

        # Step-by-step formulas
        steps = [
            ("M11", "I1 + m1*(l1/2)^2 + I2 + m2*(l1^2 + l1*l2*cos(t2) + (l2/2)^2)",
             f"{I1:.5f} + {p.m1*(p.l1/2)**2:.5f} + {I2:.5f} + {p.m2*(p.l1**2 + p.l1*p.l2*c2 + (p.l2/2)**2):.5f}",
             round(M11, 4)),
            ("M12", "I2 + m2*(l2/2)*(l1*cos(t2) + l2/2)",
             f"{I2:.5f} + {p.m2*(p.l2/2)*(p.l1*c2 + p.l2/2):.5f}",
             round(M12, 4)),
            ("M21", "= M12 (symmetric)",
             f"{round(M12, 4)}",
             round(M12, 4)),
            ("M22", "I2 + m2*(l2/2)^2",
             f"{I2:.5f} + {p.m2*(p.l2/2)**2:.5f}",
             round(M22, 4)),
        ]

        headers = ["Element", "Formula", "Substituted", "Value"]
        for ci, h in enumerate(headers, 1):
            style_header(ws.cell(row, ci), h, fill=SUBHDR_FILL, font=SUBHDR_FONT)
        for ci in range(5, 8):
            ws.cell(row, ci).value = ""
        row += 1

        for j, (elem, formula, sub, val) in enumerate(steps):
            fill = ALT_FILL if j % 2 else None
            ws.cell(row, 1).value = elem; ws.cell(row, 1).font = Font(bold=True, size=10)
            ws.cell(row, 2).value = formula
            ws.cell(row, 3).value = sub
            ws.cell(row, 4).value = val; ws.cell(row, 4).fill = GREEN_FILL
            for ci in range(1,5):
                ws.cell(row, ci).border = THIN_BORDER
                if fill and ci != 4:
                    ws.cell(row, ci).fill = fill
            row += 1

        # Matrix display
        ws.cell(row, 1).value = "M(theta) ="
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = f"[[{M11:.4f}, {M12:.4f}], [{M12:.4f}, {M22:.4f}]]"
        ws.cell(row, 2).fill = GREEN_FILL; ws.cell(row, 2).font = Font(bold=True)
        ws.merge_cells(f"B{row}:G{row}")
        row += 2

    set_col_widths(ws, {"A":12,"B":55,"C":45,"D":12,"E":12,"F":12,"G":12})


# ---- Sheet 3: GravityTorques ----

def sheet_gravity_torques(wb):
    ws = wb.create_sheet("GravityTorques")
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Gravity Torques G(theta) — Computation at 3 Configurations"
    c.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    p = ArmParams()
    arm = ArmDynamics(p)

    configs = [
        (0.0, 0.0, "theta=[0, 0]"),
        (math.pi/4, math.pi/4, "theta=[pi/4, pi/4]"),
        (math.pi/2, 0.0, "theta=[pi/2, 0]"),
    ]

    headers = ["Config", "theta1", "theta2", "G1 formula", "G1 value", "G2 formula", "G2 value"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(3, ci), h)

    for i, (t1, t2, name) in enumerate(configs):
        G = arm.gravity_torques(t1, t2)
        s1 = math.sin(t1); s12 = math.sin(t1+t2)
        g1_formula = f"(m1*l1/2 + m2*l1)*g*sin({t1:.4f}) + m2*(l2/2)*g*sin({t1+t2:.4f})"
        g2_formula = f"m2*(l2/2)*g*sin({t1+t2:.4f})"
        row = 4 + i
        fill = ALT_FILL if i % 2 else None
        data = [name, round(t1,4), round(t2,4), g1_formula, round(G[0],4), g2_formula, round(G[1],4)]
        for ci, val in enumerate(data, 1):
            style_body(ws.cell(row, ci), val, fill=fill)
        ws.cell(row, 5).fill = GREEN_FILL
        ws.cell(row, 7).fill = GREEN_FILL

    # Note
    ws["A8"].value = ("G1 = (m1*l1/2 + m2*l1)*g*sin(theta1) + m2*(l2/2)*g*sin(theta1+theta2)   "
                      "G2 = m2*(l2/2)*g*sin(theta1+theta2)")
    ws["A8"].font = NOTE_FONT; ws["A8"].fill = NOTE_FILL
    ws.merge_cells("A8:G8")

    set_col_widths(ws, {"A":20,"B":12,"C":12,"D":55,"E":12,"F":45,"G":12})


# ---- Sheet 4: MRTA_Mapping ----

def sheet_mrta_mapping(wb, prob):
    ws = wb.create_sheet("MRTA_Mapping")
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "MRTA to SNN Mapping: Each Coalition Node = LIF Neuron"
    c.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    ws["A2"].value = ("Each coalition node i maps to neuron i. "
                      "External drive I_ext_i = utility_i. "
                      "Inhibitory weight W_ij = -2.0 for conflict edges. "
                      "Winning allocation = neurons with highest spike count forming independent set.")
    ws["A2"].font = NOTE_FONT; ws["A2"].fill = NOTE_FILL
    ws.merge_cells("A2:G2")

    headers = ["Neuron#", "Coalition Label", "Task", "Utility (I_ext)", "Connections (conflict)", "W_ij", "Role"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(4, ci), h)

    for i, node in enumerate(prob.nodes):
        row = 5 + i
        conflicts = [j for j in prob.adjacency[i]]
        conflict_labels = [prob.nodes[j].label for j in conflicts]
        fill = ALT_FILL if i % 2 else None
        data = [
            i,
            node.label,
            f"t{node.task_id+1}",
            round(node.utility, 4),
            str(conflict_labels),
            -2.0 if conflicts else 0.0,
            "Optimal" if i in [0, 4] else "Suboptimal"
        ]
        for ci, val in enumerate(data, 1):
            style_body(ws.cell(row, ci), val, fill=fill)
        if i in [0, 4]:
            ws.cell(row, 7).fill = GREEN_FILL

    set_col_widths(ws, {"A":10,"B":18,"C":8,"D":18,"E":60,"F":10,"G":12})


# ---- Sheet 5: SNN_Dynamics ----

def sheet_snn_dynamics(wb, prob):
    ws = wb.create_sheet("SNN_Dynamics")
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "SNN LIF Dynamics — Voltage Traces for All 7 Neurons (200ms simulation)"
    c.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    utilities = [n.utility for n in prob.nodes]
    cfg = SNNConfig(sim_time_ms=200.0, dt_ms=0.1, restarts=1, seed=42, noise_amp=0.05)
    solver = SNNSolver(cfg)
    sim = solver.simulate(utilities, prob.adjacency, prob.lambda_penalty, record_traces=True)

    # Record at specific time points
    sample_times_ms = list(range(0, 101, 5))  # 0, 5, 10, ..., 100 ms
    dt = cfg.dt_ms
    n = prob.node_count

    headers = ["t (ms)"] + [f"V{i} ({prob.nodes[i].label})" for i in range(n)]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(3, ci), h)

    row = 4
    time_axis = sim.time_axis_ms
    voltage_traces = sim.voltage_traces

    for t_ms in sample_times_ms:
        # Find closest recorded index
        if len(time_axis) == 0:
            continue
        idx = min(range(len(time_axis)), key=lambda k: abs(time_axis[k] - t_ms))

        data = [round(t_ms, 1)] + [
            round(voltage_traces[i][idx], 4) if idx < len(voltage_traces[i]) else 0.0
            for i in range(n)
        ]
        fill = ALT_FILL if (row - 4) % 2 else None
        for ci, val in enumerate(data, 1):
            style_body(ws.cell(row, ci), val, fill=fill)
        row += 1

    # Spike times section
    row += 1
    ws.merge_cells(f"A{row}:J{row}")
    ws.cell(row, 1).value = "Spike Times per Neuron"
    ws.cell(row, 1).font = Font(bold=True, size=11)
    row += 1

    spike_headers = ["Neuron#", "Label", "Spike Times (ms)", "Spike Count"]
    for ci, h in enumerate(spike_headers, 1):
        style_header(ws.cell(row, ci), h)
    row += 1

    for i, sr in enumerate(sim.spike_records):
        fill = ALT_FILL if i % 2 else None
        spk_str = str([round(t, 1) for t in sr.spike_times_ms[:20]])
        data = [i, prob.nodes[i].label, spk_str, sr.spike_count]
        for ci, val in enumerate(data, 1):
            style_body(ws.cell(row, ci), val, fill=fill)
        if i in [0, 4]:
            ws.cell(row, 4).fill = GREEN_FILL
        row += 1

    ws.column_dimensions["A"].width = 10
    for ci in range(2, n+2):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.column_dimensions[get_column_letter(n+2)].width = 14
    ws.column_dimensions[get_column_letter(n+3)].width = 50
    ws.column_dimensions[get_column_letter(n+4)].width = 14


# ---- Sheet 6: SpikeCounts ----

def sheet_spike_counts(wb, prob):
    ws = wb.create_sheet("SpikeCounts")
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "SNN Spike Counts and Final Allocation"
    c.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    utilities = [n.utility for n in prob.nodes]
    cfg = SNNConfig(sim_time_ms=200.0, dt_ms=0.1, restarts=5, seed=42)
    solver = SNNSolver(cfg)
    result = solver.solve(utilities, prob.adjacency, prob.lambda_penalty)

    headers = ["Neuron#", "Label", "Utility (I_ext)", "Spike Count", "Selected?", "In Optimal?"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(3, ci), h)

    for i in range(prob.node_count):
        row = 4 + i
        is_selected = i in result.selected
        is_optimal = i in [0, 4]
        fill = GREEN_FILL if is_selected else ALT_FILL if i % 2 else None
        data = [
            i, prob.nodes[i].label, round(utilities[i], 4),
            result.spike_counts[i] if i < len(result.spike_counts) else 0,
            "YES" if is_selected else "no",
            "YES" if is_optimal else "no",
        ]
        for ci, val in enumerate(data, 1):
            style_body(ws.cell(row, ci), val, fill=fill)

    summary_row = 4 + prob.node_count + 1
    ws.cell(summary_row, 1).value = "SNN Result"
    ws.cell(summary_row, 1).font = Font(bold=True)
    ws.cell(summary_row, 2).value = str([prob.nodes[i].label for i in result.selected])
    ws.cell(summary_row, 3).value = f"Utility = {result.utility:.4f}"
    ws.cell(summary_row, 4).value = f"Feasible = {result.feasible}"
    ws.cell(summary_row, 5).value = f"Runtime = {result.runtime_ms:.2f} ms"
    for ci in range(1, 6):
        ws.cell(summary_row, ci).fill = GREEN_FILL
        ws.cell(summary_row, ci).font = Font(bold=True)

    set_col_widths(ws, {"A":10,"B":20,"C":16,"D":14,"E":12,"F":12})


# ---- Sheet 7: Comparison ----

def sheet_comparison(wb, prob):
    ws = wb.create_sheet("Comparison")
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "SNN vs OIM Comparison on 3R2T Instance"
    c.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    utilities = [n.utility for n in prob.nodes]

    # OIM
    oim_r = solve_kuramoto_oim(prob, seed=42)

    # SNN
    cfg = SNNConfig(sim_time_ms=200.0, dt_ms=0.1, restarts=5, seed=42)
    solver = SNNSolver(cfg)
    snn_r = solver.solve(utilities, prob.adjacency, prob.lambda_penalty)

    headers = ["Solver", "Selected Nodes", "Utility", "Feasible", "Runtime (ms)", "Finds Optimal?"]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(3, ci), h)

    optimal_utility = 9.1787
    rows_data = [
        ("OIM (Kuramoto)", [prob.nodes[i].label for i in oim_r.selected],
         oim_r.utility, oim_r.feasible, oim_r.runtime_ms),
        ("SNN (LIF Neurons)", [prob.nodes[i].label for i in snn_r.selected],
         snn_r.utility, snn_r.feasible, snn_r.runtime_ms),
    ]

    for i, (name, labels, util, feasible, rtime) in enumerate(rows_data):
        row = 4 + i
        is_opt = abs(util - optimal_utility) < 0.001
        fill = GREEN_FILL if is_opt else ALT_FILL
        data = [name, str(labels), round(util, 4), str(feasible), round(rtime, 2), "YES" if is_opt else "NO"]
        for ci, val in enumerate(data, 1):
            style_body(ws.cell(row, ci), val, fill=fill)

    # Notes
    ws["A7"].value = ("Both OIM and SNN solve the same MWIS/QUBO problem. "
                      "OIM uses Kuramoto oscillator dynamics (analog physics). "
                      "SNN uses LIF neuron spiking dynamics (neuromorphic). "
                      "Optimal = {r3}->t1 + {r1}->t2, utility = 9.1787.")
    ws["A7"].font = NOTE_FONT; ws["A7"].fill = NOTE_FILL
    ws.merge_cells("A7:F7")

    set_col_widths(ws, {"A":22,"B":50,"C":12,"D":10,"E":16,"F":14})


# ---- Main ----

def main():
    instance = get_instance()
    prob = build_mwis_problem(instance, coalition_bound=2, lambda_penalty=8.0)

    wb = openpyxl.Workbook()
    del wb["Sheet"]

    sheet_arm_params(wb)
    sheet_inertia_matrix(wb)
    sheet_gravity_torques(wb)
    sheet_mrta_mapping(wb, prob)
    sheet_snn_dynamics(wb, prob)
    sheet_spike_counts(wb, prob)
    sheet_comparison(wb, prob)

    out_path = Path(__file__).parent / "snn_2dof_dataset.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


if __name__ == "__main__":
    main()
